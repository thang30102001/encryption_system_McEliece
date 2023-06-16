import openpyxl
import os
import joblib
import shutil
import numpy as np


class Excel_input:
    def __init__(self, new_file_name_finally, path_save_read, choose_type_file, already_assigned, encoder, t, k, n, orther, private_key, direc_cipherText, data_line):
        self.t = t
        self.k = k
        self.n = n
        self.path_save_read = path_save_read
        self.choose_type_file = choose_type_file
        self.encoder = encoder
        self.already_assigned = already_assigned
        self.P_inv = orther['P_inv']
        self.S_inv = orther['S_inv']
        self.S = private_key['S']
        self.ciphertext = direc_cipherText['ciphertext']
        self.data_line = data_line
        self.path_location_save = None
        self.encoder_path = None
        self.new_file_name_finally = new_file_name_finally

    def Excel_Write_input(self):
        path_cipher = os.path.join(self.path_save_read, "CipherText")
        if not os.path.exists(path_cipher):
            os.mkdir(path_cipher)

        file_encoder_name = "encoder.joblib"
        self.encoder_path = os.path.join(path_cipher, file_encoder_name)
        joblib.dump(self.encoder, self.encoder_path)
        print(f"\t- Save encoder object    in: {self.encoder_path}")

        workbook = openpyxl.Workbook()
        sheet_t = workbook.active
        sheet_t.title = "t"
        sheet_k = workbook.create_sheet("k")
        sheet_n = workbook.create_sheet("n")
        sheet_S = workbook.create_sheet("S")
        sheet_S_inv = workbook.create_sheet("S_inv")
        sheet_P_inv = workbook.create_sheet("P_inv")
        sheet_already_assigned = workbook.create_sheet("already_assigned")
        file_name = "Matrix.xlsx"
        data_path = os.path.join(path_cipher, file_name)
        workbook.save(data_path)

        cell_t = sheet_t.cell(row=1, column=1)
        cell_t.value = self.t
        print(f"\t- Save t                 in: {data_path}")

        cell_k = sheet_k.cell(row=1, column=1)
        cell_k.value = self.k
        print(f"\t- Save k                 in: {data_path}")

        cell_n = sheet_n.cell(row=1, column=1)
        cell_n.value = self.k
        print(f"\t- Save n                 in: {data_path}")

        for i in range(self.S.shape[0]):
            for j in range(self.S.shape[1]):
                cell_S = sheet_S.cell(row=i+1, column=j+1)
                cell_S.value = self.S[i][j]
        print(f"\t- Save matrix S          in: {data_path}")

        for i in range(self.S_inv.shape[0]):
            for j in range(self.S_inv.shape[1]):
                cell_S_inv = sheet_S_inv.cell(row=i+1, column=j+1)
                cell_S_inv.value = self.S_inv[i][j]
        print(f"\t- Save matrix S^-1       in: {data_path}")

        for i in range(self.P_inv.shape[0]):
            for j in range(self.P_inv.shape[1]):
                cell_P_inv = sheet_P_inv.cell(row=i+1, column=j+1)
                cell_P_inv.value = self.P_inv[i][j]
        print(f"\t- Save matrix P^-1       in: {data_path}")

        for i in range(self.already_assigned.shape[0]):
            for j in range(self.already_assigned.shape[1]):
                cell_sheet_already_assigned = sheet_already_assigned.cell(row=i+1, column=j+1)
                cell_sheet_already_assigned.value = self.already_assigned[i][j]
        print(f"\t- Save already_assigned  in: {data_path}")

        workbook.save(data_path)
        to_path = os.path.join(os.getcwd(), "..", "data")

        if self.choose_type_file:
            txt_file = self.new_file_name_finally
            txt_path = os.path.join(path_cipher, txt_file)
            with open(txt_path, "a", encoding='utf-8') as f:
                np.savetxt(f, self.ciphertext, fmt="%s", delimiter=", ")

            print(f"\t- Save cipher            in: {txt_path}")

            self.path_location_save = txt_path
            return self.path_location_save, self.encoder_path

        else:
            workbook_ex = openpyxl.Workbook()
            sheet_cipher = workbook_ex.active
            sheet_cipher.title = "cipher"
            file_name_cipher = self.new_file_name_finally
            data_path_cipher = os.path.join(path_cipher, file_name_cipher)
            workbook_ex.save(data_path_cipher)

            for i in range(self.ciphertext.shape[0]):
                for j in range(self.ciphertext.shape[1]):
                    cell_cipher = sheet_cipher.cell(row=i+1, column=j+1)
                    cell_cipher.value = self.ciphertext[i][j]

            print(f"\t- Save cipher            in: {data_path_cipher}\n")
            workbook_ex.save(data_path_cipher)
            
            self.path_location_save = data_path_cipher
            return self.path_location_save, self.encoder_path


            